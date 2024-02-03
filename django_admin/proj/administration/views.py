from datetime import datetime
import os
import re

from django.shortcuts import render, redirect, get_object_or_404
from .forms import ExcelForm, PriceListForm
import openpyxl as op
from glob import glob

from .models import Item, InStock, TelegramIdImage


def main(request):
    return render(request, 'administration/main_page.html')

def excel_confirm(request):
    return render(request, 'administration/excel_confirm.html')

def price_list_confirm(request):
    return render(request, 'administration/price_list_confirm.html')

def excel_input(request):
    error = None
    file = ExcelForm(request.POST, request.FILES)

    class EmptyFirstCell(Exception):
        pass

    class ALotOfRows(Exception):
        pass

    if request.method == "POST":  # данные отправляются из формы
        if file.is_valid():
            try:
                # получение данных из файла
                check_files = glob(f"media/excel/*")
                if len(check_files) == 0:
                    file.save()
                else:
                    for item in check_files:
                        os.remove(item)
                    file.save()

                tables = glob(f"media/excel/*")
                table = tables[0]

                wb = op.load_workbook(table)
                ws = wb[wb.sheetnames[0]]

                if ws.cell(row=1, column=1).value is None:
                    raise EmptyFirstCell()

                # запись данных в бд

                stop = False
                row = 1

                cols = {
                    'cat': 1,
                    'name': 2,
                    'height': 3,
                    'date': 4,
                    'count': 5,
                    'mult': 6,
                    'price': 7
                }

                category = ""
                write: bool = False

                while not stop:
                    item = {
                        'name': "",
                        'desc': "",
                        'price': 0,
                        'count': 0,
                        'mult': 0,
                        'date': ""
                    }

                    row_cells = [1, 1, 1, 1, 1, 1, 1]

                    for col in range(1, 8):
                        cur_cell = ws.cell(row=row, column=col).value
                        next_cell_y = ws.cell(row=row + 1, column=col).value
                        next_cell_x = ws.cell(row=row, column=col + 1).value

                        if row_cells[0] is None and row_cells[1] is None and category != "":
                            stop = True
                            write = False
                            break
                        else:
                            if cur_cell is not None:
                                if next_cell_x is None and next_cell_y is None and col == cols['cat']:
                                    cat = re.split(r'^[^а-яА-ЯёЁ]+', f'{cur_cell}')
                                    if cat[0] != '':
                                        category = cat[0]
                                    else:
                                        category = cat[1]
                                else:
                                    if col == cols['name']:
                                        item['name'] = cur_cell
                                    elif col == cols['height']:
                                        item['desc'] = cur_cell
                                    elif col == cols['date']:
                                        item['date'] = cur_cell
                                    elif col == cols['count']:
                                        item['count'] = cur_cell
                                    elif col == cols['mult']:
                                        item['mult'] = cur_cell
                                    elif col == cols['price']:
                                        item['price'] = cur_cell
                                    write = True
                            else:
                                write = False

                        row_cells[col-1] = cur_cell

                    if write and category != '':
                        try:
                            date = datetime.strptime(item['date'], '%d.%m.%Y')

                            if TelegramIdImage.objects.filter(name__icontains=item['name']):
                                row_db = get_object_or_404(TelegramIdImage, name=item['name'])
                                tg_image_id = row_db.id_image
                                if datetime.today() < date:
                                    item_otw = Item(name=item['name'], description=f"{item['desc']}",
                                                    category=category, price=item['price'], col=item['count'],
                                                    mult=item['mult'], item_date=date.strftime('%Y-%m-%d'),
                                                    image_id=tg_image_id)
                                    item_otw.save()
                                else:
                                    item_inst = InStock(name=item['name'], description=f"{item['desc']}",
                                                        category=category, price=item['price'], col=item['count'],
                                                        mult=item['mult'], item_date=date.strftime('%Y-%m-%d'),
                                                        image_id=tg_image_id)
                                    item_inst.save()
                            else:
                                if datetime.today() < date:
                                    item_otw = Item(name=item['name'], description=f"{item['desc']}",
                                                    category=category, price=item['price'], col=item['count'],
                                                    mult=item['mult'], item_date=date.strftime('%Y-%m-%d'))
                                    item_otw.save()
                                else:
                                    item_inst = InStock(name=item['name'], description=f"{item['desc']}",
                                                        category=category, price=item['price'], col=item['count'],
                                                        mult=item['mult'], item_date=date.strftime('%Y-%m-%d'))
                                    item_inst.save()
                        except ValueError:
                            os.remove(table)

                            error = "Произошла ошибка при обработке файла! Загрузите таблицу нужного формата"

                            form = ExcelForm()

                            data = {
                                'form': form,
                                'error': error
                            }

                            return render(request, 'administration/excel_input.html', data)

                    row += 1

                    if row > 500:
                        raise ALotOfRows()

                os.remove(table)

                return redirect("/excelconfirm") #при добавлении, перенаправление на страницу

            except (AttributeError, EmptyFirstCell, ALotOfRows):
                error = "Произошла ошибка при обработке файла! Загрузите таблицу нужного формата"

                form = ExcelForm()

                data = {
                    'form': form,
                    'error': error
                }

                return render(request, 'administration/excel_input.html', data)
        else:
            error = "Неподходящий формат! Выберите файл с расширением xls или xlsx"

    form = ExcelForm()

    data = {
        'form': form,
        'error': error
    }

    return render(request, 'administration/excel_input.html', data)

def price_list(request):
    error = None
    file = PriceListForm(request.POST, request.FILES)

    if request.method == "POST":
        if file.is_valid():
            if len(os.listdir('media/pricelist')) == 0:
                file.save()
            else:
                content = glob(f"media/pricelist/*")
                for item in content:
                    os.remove(item)
                file.save()

            return redirect("/pricelistconfirm")
        else:
            error = "Произошла ошибка!"

    form = PriceListForm()

    data = {
        'form': form,
        'error': error
    }

    return render(request, 'administration/price_list.html', data)
